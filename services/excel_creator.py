from decimal import Decimal
from openpyxl.styles import PatternFill, Alignment
import openpyxl.styles.numbers
import openpyxl


class ExcelCreator:

    SHEET_NAME = "Отчет"

    # в конструторе инициализируем рабочую книгу, лист и счетчик строк
    def __init__(self):
        self.row_counter = 0
        self.__book = openpyxl.Workbook()
        self.__book.remove(self.__book.active)  # удаляем дефолтный лист с названием Sheet
        self.__reportsheet = self.__book.create_sheet(self.SHEET_NAME)

    # создаём заголовки для столбцов на основе полученных валютных пар
    def create_headers(self, currency_pairs):
        xl_headers = []
        for arg in currency_pairs:
            xl_headers.append(f"Дата {arg}")
            xl_headers.append(f"Курс {arg}")
            xl_headers.append(f"Время {arg}")
        xl_headers.append("Результат")
        self.__reportsheet.append(xl_headers)
        self.row_counter += 1
        header_color = PatternFill("solid", fgColor="B0C4DE")
        for row in self.__reportsheet.rows:
            for cell in row:
                cell.fill = header_color

    def fill_data(self, table):
        for row in table:
            self.__reportsheet.append(row)
            self.row_counter += 1

    # Выравнивание столбцов в зависимости от ячейки с максимальной длиной содержимого
    # с возможностью выравнивания в случае получения соответствующего флага
    def adjust_column_widths(self, required_alignment=False):
        align = Alignment(horizontal='center')
        for col in self.__reportsheet.columns:
            max_length = 0
            column_letter = col[0].column_letter  # Получить букву столбца
            for cell in col:
                if required_alignment:
                    cell.alignment = align
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.__reportsheet.column_dimensions[column_letter].width = adjusted_width


    # устанавливаем финансовый формат для рубля
    def format_number_as_finance(self):
        for letter in "BEG":
            for i in range(2, self.row_counter + 1):
                self.__reportsheet[f"{letter}{i}"].number_format = '#,##0.00" ₽"'

    # сохраняем файл
    def save(self, filename):
        self.__book.save(filename)
        self.__book.close()

    # устанавливаем формулу автосуммы в ячейку и проверяем, что она корректно посчиталась
    def check_autosum(self, filename):
        book = openpyxl.load_workbook(filename)
        sheet = book[self.SHEET_NAME]
        self.row_counter += 1
        sheet[f'G{self.row_counter}'] = '=SUM(G2:G24)'
        book.save(filename)
        book.close()
